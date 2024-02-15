from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, colors, NamedStyle, PatternFill
from app.business_logic.utilities import static_data as sd
import os


def create_sheet_for_print(sheet, rows, main_header, category_header):

    # adding the header to sheet
    sheet.append(sd.CLUBBED_OUTPUT_HEADER)
    row = sheet[1]
    for cell in row:
        cell.style = main_header
    # Write the data rows
    row_num = 2
    for r in rows:
        sheet.append(list(r.values()))
        if r[sd.CONSEQUENCE_COLUMN] == " ":
            for cell in sheet[row_num]:
                cell.style = category_header
        row_num += 1

# temp function for system files
def create_sheet_for_print_temp(sheet, rows, main_header, category_header):

    # adding the header to sheet
    sheet.append(sd.REGULAR_OUTPUT_HEADER)
    row = sheet[1]
    for cell in row:
        cell.style = main_header
    # Write the data rows
    row_num = 2
    for r in rows:
        sheet.append(list(r.values()))
        if r[sd.CONSEQUENCE_COLUMN] == " ":
            for cell in sheet[row_num]:
                cell.style = category_header
        row_num += 1


def create_sheet_regular(sheet, rows, main_header, category_header, score_color):

    # adding header to the sheet
    sheet.append(sd.CLUBBED_OUTPUT_HEADER)
    row = sheet[1]
    for cell in row:
        cell.style = main_header
    row = 2
    for r in rows:
        # 8, 9, 10 and 11 should be blue if empty
        sheet.append(list(r.values()))
        sheet.cell(row=row, column=2).style = score_color
        sheet.cell(row=row, column=4).style = score_color
        sheet.cell(row=row, column=6).style = score_color
        sheet.cell(row=row, column=13).style = score_color
        sheet.cell(row=row, column=15).style = score_color
        sheet.cell(row=row, column=26).style = score_color
        sheet.cell(row=row, column=28).style = score_color
        if r["Clinical Consequence"] == "NA":
            sheet.cell(row=row, column=8).style = category_header
            sheet.cell(row=row, column=9).style = category_header
            sheet.cell(row=row, column=10).style = category_header
            sheet.cell(row=row, column=11).style = category_header

        row += 1

def create_sheet_transcell(sheet, rows, main_header, category_header, score_color):

    # adding header to the sheet
    sheet.append(sd.TRANSCELL_OUTPUT_HEADER)
    row = sheet[1]
    for cell in row:
        cell.style = main_header
    row = 2
    for r in rows:
        # 8, 9, 10 and 11 should be blue if empty
        sheet.append(list(r.values()))
        sheet.cell(row=row, column=2).style = score_color
        sheet.cell(row=row, column=4).style = score_color
        sheet.cell(row=row, column=6).style = score_color
        sheet.cell(row=row, column=13).style = score_color
        sheet.cell(row=row, column=15).style = score_color
        sheet.cell(row=row, column=26).style = score_color
        sheet.cell(row=row, column=28).style = score_color
        row += 1

def main_header_style():
    main_header = NamedStyle(name="main_header")
    main_header.font = Font(bold=True)
    main_header.fill = PatternFill(fgColor="ffcc66", fill_type="solid")
    main_header.border = Border(bottom=Side(border_style="thick"))
    main_header.alignment = Alignment(horizontal="center")
    return main_header


def category_header_style():
    category_header = NamedStyle(name="category_header")
    category_header.fill = PatternFill(fgColor="99ccff", fill_type="solid")
    category_header.font = Font(bold=True)
    category_header.alignment = Alignment(horizontal="center")
    return category_header


def score_cell_color():
    score_cells = NamedStyle(name="score_cells")
    score_cells.fill = PatternFill(fgColor="5cd65c", fill_type="solid")
    score_cells.font = Font(bold=True)
    score_cells.alignment = Alignment(horizontal="center")
    return score_cells


def write_indels(rows, sample):
    # check if output directory exists
    if os.path.isdir(sd.OUTPUT_DIR):
        if os.path.isdir(f"{sd.OUTPUT_DIR}/{sample}"):
            print("directory exists")
        else:
            try:
                os.makedirs(f"{sd.OUTPUT_DIR}/{sample}", exist_ok=True)
            except OSError as oe:
                print("Directory cannot be created :", oe)
    else:
        # if not create the directory
        try:
            os.makedirs(f"{sd.OUTPUT_DIR}/{sample}", exist_ok=True)
            print("Directory Created Successully: ", sample)
        except OSError as oe:
            print("Directory cannot be created :", oe)
    wb = Workbook()
    sheet_indel = wb.active
    sheet_indel.title = sd.INDEL_SHEET_NAME
    indel_header = NamedStyle(name="indel_header")
    indel_header.font = Font(bold=True)
    indel_header.fill = PatternFill(fgColor="ffcc66", fill_type="solid")
    indel_header.border = Border(bottom=Side(border_style="thick"))
    indel_header.alignment = Alignment(horizontal="center")
    # write to side effects sheet
    # write header
    h_list = list(rows[0].keys())
    sheet_indel.append(h_list)

    row = sheet_indel[1]
    for cell in row:
        cell.style = indel_header
    # write data rows
    row_num = 2
    for r in rows:
        sheet_indel.append(list(r.values()))
        row_num += 1
    wb.save(f"{sd.OUTPUT_DIR}/{sample}/{sample}{sd.INDEL_OUTPUT_SUFFIX}")


def write_pre_processed_vcf(rows, sample):
    if os.path.isdir(sd.OUTPUT_DIR):
        if os.path.isdir(f"{sd.OUTPUT_DIR}/{sample}"):
            print("directory exists")
        else:
            try:
                os.makedirs(f"{sd.OUTPUT_DIR}/{sample}", exist_ok=True)
            except OSError as oe:
                print("Directory cannot be created :", oe)
    else:
        # if not create the directory
        try:
            os.makedirs(f"{sd.OUTPUT_DIR}/{sample}", exist_ok=True)
            print("Directory Created Successully: ", sample)
        except OSError as oe:
            print("Directory cannot be created :", oe)
    wb = Workbook()
    vcf_sheet = wb.active
    vcf_sheet.title = "VCF_processed"
    vcf_header_style = NamedStyle(name="vcf_header")
    vcf_header_style.font = Font(bold=True)
    vcf_header_style.fill = PatternFill(fgColor="ffcc66", fill_type="solid")
    vcf_header_style.border = Border(bottom=Side(border_style="thick"))
    vcf_header_style.alignment = Alignment(horizontal="center")
    h_list = list(rows[0].keys())
    vcf_sheet.append(h_list)

    row = vcf_sheet[1]
    for cell in row:
        cell.style = vcf_header_style
    # write data rows
    row_num = 2
    for r in rows:
        vcf_sheet.append(list(r.values()))
        row_num += 1
    wb.save(f"{sd.OUTPUT_DIR}/{sample}/{sample}{sd.VCF_PROCESSED_FILE}")
